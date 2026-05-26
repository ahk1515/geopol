# =============================================================
# GÉOPOL — Parser Energy Institute Statistical Review
# Source : Energy Institute — Statistical Review of World Energy
# https://www.energyinst.org/statistical-review
#
# Schéma identite :
#   indicator     = energie_production | energie_reserves
#                   mineraux_production | mineraux_reserves
#   subcategory   = petrole | gaz | charbon | nucleaire |
#                   solaire | eolien | hydro | geo_biomasse |
#                   renouvelables_total | cobalt | lithium |
#                   cuivre | nickel | zinc | bauxite | aluminium |
#                   graphite | terres_rares | manganese | etain |
#                   vanadium | pgm
#
# Mode : semi-automatique
#   → déposer le fichier .xlsx dans etl/sources/uploads/
#   → nom attendu : EI-Stats-Review-ALL-data.xlsx (ou similaire)
# =============================================================

import sqlite3
import os
import sys
import re

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from etl.config import ANNEE_DEBUT, ANNEE_FIN, PATH_DB

SOURCE = "Energy Institute"

# -------------------------------------------------------------
# CORRESPONDANCE NOM EI → ISO3
# EI utilise des noms anglais parfois non standard
# -------------------------------------------------------------
EI_TO_ISO3 = {
    "Algeria": "DZA", "Angola": "AGO", "Argentina": "ARG",
    "Australia": "AUS", "Austria": "AUT", "Azerbaijan": "AZE",
    "Bahrain": "BHR", "Bangladesh": "BGD", "Belarus": "BLR",
    "Belgium": "BEL", "Bolivia": "BOL", "Brazil": "BRA",
    "Brunei": "BRN", "Bulgaria": "BGR", "Cameroon": "CMR",
    "Canada": "CAN", "Chile": "CHL", "China": "CHN",
    "Colombia": "COL", "Congo": "COG", "Croatia": "HRV",
    "Cuba": "CUB", "Cyprus": "CYP", "Czechia": "CZE",
    "Czech Republic": "CZE", "Denmark": "DNK", "Ecuador": "ECU",
    "Egypt": "EGY", "Estonia": "EST", "Ethiopia": "ETH",
    "Finland": "FIN", "France": "FRA", "Gabon": "GAB",
    "Germany": "DEU", "Ghana": "GHA", "Greece": "GRC",
    "Guatemala": "GTM", "Hungary": "HUN", "India": "IND",
    "Indonesia": "IDN", "Iran": "IRN", "Iraq": "IRQ",
    "Ireland": "IRL", "Israel": "ISR", "Italy": "ITA",
    "Japan": "JPN", "Jordan": "JOR", "Kazakhstan": "KAZ",
    "Kenya": "KEN", "Kuwait": "KWT", "Kyrgyzstan": "KGZ",
    "Latvia": "LVA", "Libya": "LBY", "Lithuania": "LTU",
    "Madagascar": "MDG", "Malaysia": "MYS", "Mexico": "MEX",
    "Morocco": "MAR", "Mozambique": "MOZ", "Myanmar": "MMR",
    "Netherlands": "NLD", "New Caledonia": "NCL",
    "New Zealand": "NZL", "Nigeria": "NGA", "Norway": "NOR",
    "Oman": "OMN", "Pakistan": "PAK", "Papua New Guinea": "PNG",
    "Peru": "PER", "Philippines": "PHL", "Poland": "POL",
    "Portugal": "PRT", "Qatar": "QAT", "Romania": "ROU",
    "Russia": "RUS", "Russian Federation": "RUS",
    "Saudi Arabia": "SAU", "Serbia": "SRB", "Singapore": "SGP",
    "Slovakia": "SVK", "Slovenia": "SVN", "South Africa": "ZAF",
    "South Korea": "KOR", "Spain": "ESP", "Sri Lanka": "LKA",
    "Sudan": "SDN", "Sweden": "SWE", "Switzerland": "CHE",
    "Taiwan": "TWN", "Thailand": "THA", "Trinidad & Tobago": "TTO",
    "Tunisia": "TUN", "Turkey": "TUR", "Turkiye": "TUR",
    "Turkmenistan": "TKM", "UAE": "ARE",
    "United Arab Emirates": "ARE", "UK": "GBR",
    "United Kingdom": "GBR", "US": "USA", "USA": "USA",
    "United States": "USA", "Ukraine": "UKR",
    "Uzbekistan": "UZB", "Venezuela": "VEN", "Vietnam": "VNM",
    "Viet Nam": "VNM", "Zambia": "ZMB", "Zimbabwe": "ZWE",
    "DR Congo": "COD", "Congo, Dem. Rep.": "COD",
    "Bolivia (Plurinational State of)": "BOL",
    "Iran, Islamic Republic of": "IRN",
    "Korea, Republic of": "KOR",
}

# Mots-clés pour exclure les lignes non-pays (totaux, régions, notes)
EXCLUDE_KEYWORDS = [
    "total", "other", "rest of", "of which", "oecd", "opec",
    "non-oecd", "non-opec", "european union", "source", "note",
    "includes", "excludes", "less than", "n/a", "^", "♦", "#",
    "ussr", "cis", "world", "region",
]

def is_country_row(name):
    """Retourne True si la ligne correspond à un pays (pas un total/région/note)."""
    if not name or not isinstance(name, str):
        return False
    name_l = name.strip().lower()
    if not name_l:
        return False
    for kw in EXCLUDE_KEYWORDS:
        if kw in name_l:
            return False
    # Lignes commençant par espace = sous-totaux EI
    if name.startswith("  ") or name.startswith("\t"):
        return False
    return True


def parse_value(v):
    """Parse une valeur numérique, retourne None si invalide."""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        val = float(v)
        return val if val > 0 else None
    s = str(v).strip().replace(",", "").replace("n/a", "").replace("^", "").replace("♦", "")
    # Nettoyer les notes de bas de page (chiffres en exposant collés)
    s = re.sub(r'[a-zA-Z#*†‡]+$', '', s).strip()
    # Nettoyer les noms de pays avec chiffres collés (ex: "Brazil1")
    if not s:
        return None
    try:
        val = float(s)
        return val if val > 0 else None
    except ValueError:
        return None


def clean_country_name(name):
    """Nettoie le nom du pays (chiffres et symboles en suffixe)."""
    if not name:
        return name
    return re.sub(r'[\d#*†‡\s]+$', '', str(name)).strip()


# -------------------------------------------------------------
# PARSERS GÉNÉRIQUES
# -------------------------------------------------------------

def parse_series_sheet(ws, indicator, subcategory, unit):
    """
    Parse un onglet de série temporelle (structure standard EI) :
    - Ligne 2 (idx) : années
    - Ligne 3 (idx) : vide
    - Lignes suivantes : pays + valeurs
    Retourne une liste de (iso3, year, value).
    """
    rows = list(ws.iter_rows(values_only=True))

    # Trouver la ligne d'en-tête (contient des années entières)
    header_idx = None
    for i, row in enumerate(rows[:10]):
        if any(isinstance(v, int) and 1960 <= v <= 2030 for v in row):
            header_idx = i
            break

    if header_idx is None:
        print(f"  ⚠️  En-tête non trouvé pour {subcategory}")
        return []

    # Années disponibles et leurs index de colonne
    header_row = rows[header_idx]
    year_cols = {
        col_idx: v
        for col_idx, v in enumerate(header_row)
        if isinstance(v, int) and ANNEE_DEBUT <= v <= ANNEE_FIN
    }

    if not year_cols:
        print(f"  ⏭️  Aucune année dans la plage {ANNEE_DEBUT}-{ANNEE_FIN} pour {subcategory}")
        return []

    results = []
    for row in rows[header_idx + 2:]:  # +2 pour sauter la ligne vide
        raw_name = row[0]
        name = clean_country_name(raw_name)

        if not is_country_row(name):
            continue

        iso3 = EI_TO_ISO3.get(name)
        if not iso3:
            continue

        for col_idx, year in year_cols.items():
            if col_idx >= len(row):
                continue
            val = parse_value(row[col_idx])
            if val is not None:
                results.append((iso3, year, val))

    return results


def parse_reserves_snapshot(ws, indicator, subcategory, unit):
    """
    Parse les onglets réserves pétrole/gaz (structure snapshot) :
    - Ligne 3 : labels "at end 2000", "at end 2010"...
    - Données : pays + valeurs par colonne année
    Retourne une liste de (iso3, year, value).
    """
    rows = list(ws.iter_rows(values_only=True))

    # Ligne 3 contient "at end XXXX"
    year_row = rows[3]
    year_cols = {}
    for col_idx, v in enumerate(year_row):
        if v and isinstance(v, str) and "at end" in v.lower():
            m = re.search(r'(\d{4})', v)
            if m:
                yr = int(m.group(1))
                if ANNEE_DEBUT <= yr <= ANNEE_FIN:
                    year_cols[col_idx] = yr

    if not year_cols:
        print(f"  ⏭️  Aucune année snapshot dans la plage pour {subcategory}")
        return []

    results = []
    for row in rows[7:]:  # données commencent ligne 8
        raw_name = row[0]
        name = clean_country_name(raw_name)

        if not is_country_row(name):
            continue

        iso3 = EI_TO_ISO3.get(name)
        if not iso3:
            continue

        for col_idx, year in year_cols.items():
            if col_idx >= len(row):
                continue
            val = parse_value(row[col_idx])
            if val is not None:
                results.append((iso3, year, val))

    return results


def parse_coal_reserves(ws, indicator, subcategory, unit):
    """
    Parse l'onglet Coal - Reserves (structure différente : total col 3).
    """
    rows = list(ws.iter_rows(values_only=True))

    # Pas d'années — une seule valeur (réserves actuelles, col "Total")
    # Colonne 3 = Total, ligne 5 = unité "Million tonnes"
    # On stocke en year = ANNEE_FIN comme proxy

    results = []
    for row in rows[6:]:
        raw_name = row[0]
        name = clean_country_name(raw_name)

        if not is_country_row(name):
            continue

        iso3 = EI_TO_ISO3.get(name)
        if not iso3:
            continue

        # Colonne 3 = Total (anthracite + sub-bituminous)
        val = parse_value(row[3]) if len(row) > 3 else None
        if val is not None:
            results.append((iso3, ANNEE_FIN, val))

    return results


def parse_mineral_sheet(ws, subcategory):
    """
    Parse un onglet minéral P-R (Production and Reserves).
    Structure : ligne 2 = années production, bloc réserves séparé.
    Retourne (prod_rows, res_rows) — deux listes de (iso3, year, value).
    """
    rows = list(ws.iter_rows(values_only=True))

    # Trouver la ligne d'en-tête production (contient années)
    header_idx = None
    for i, row in enumerate(rows[:10]):
        if any(isinstance(v, int) and 1960 <= v <= 2030 for v in row):
            header_idx = i
            break

    if header_idx is None:
        return [], []

    header_row = rows[header_idx]
    year_cols = {
        col_idx: v
        for col_idx, v in enumerate(header_row)
        if isinstance(v, int) and ANNEE_DEBUT <= v <= ANNEE_FIN
    }

    # Trouver le bloc réserves (chercher "reserve" dans col A)
    reserves_start = None
    for i, row in enumerate(rows[header_idx:], start=header_idx):
        if row[0] and "reserve" in str(row[0]).lower():
            reserves_start = i
            break

    # Parser production (entre header+2 et reserves_start)
    prod_end = reserves_start if reserves_start else len(rows)
    prod_rows = []
    for row in rows[header_idx + 2: prod_end]:
        raw_name = row[0]
        name = clean_country_name(raw_name)
        if not is_country_row(name):
            continue
        iso3 = EI_TO_ISO3.get(name)
        if not iso3:
            continue
        for col_idx, year in year_cols.items():
            if col_idx >= len(row):
                continue
            val = parse_value(row[col_idx])
            if val is not None:
                prod_rows.append((iso3, year, val))

    # Parser réserves (dernière colonne disponible = réserves récentes)
    res_rows = []
    if reserves_start:
        # Trouver la ligne d'en-tête réserves (années ou "Reserves XXXX")
        res_header = None
        for i, row in enumerate(rows[reserves_start:], start=reserves_start):
            if any(isinstance(v, int) and 1960 <= v <= 2030 for v in row):
                res_header = i
                break

        if res_header:
            res_year_cols = {
                col_idx: v
                for col_idx, v in enumerate(rows[res_header])
                if isinstance(v, int) and ANNEE_DEBUT <= v <= ANNEE_FIN
            }
            for row in rows[res_header + 2:]:
                raw_name = row[0]
                name = clean_country_name(raw_name)
                if not is_country_row(name):
                    continue
                iso3 = EI_TO_ISO3.get(name)
                if not iso3:
                    continue
                for col_idx, year in res_year_cols.items():
                    if col_idx >= len(row):
                        continue
                    val = parse_value(row[col_idx])
                    if val is not None:
                        res_rows.append((iso3, year, val))
        else:
            # Pas d'années dans réserves → snapshot sur dernière colonne non-vide
            for row in rows[reserves_start + 2:]:
                raw_name = row[0]
                name = clean_country_name(raw_name)
                if not is_country_row(name):
                    continue
                iso3 = EI_TO_ISO3.get(name)
                if not iso3:
                    continue
                # Dernière valeur non-None de la ligne
                vals = [parse_value(v) for v in row[1:] if parse_value(v) is not None]
                if vals:
                    res_rows.append((iso3, ANNEE_FIN, vals[-1]))

    return prod_rows, res_rows


# -------------------------------------------------------------
# INSERTION EN BASE
# -------------------------------------------------------------

def ensure_identite_table(conn):
    conn.execute("""
        CREATE TABLE IF NOT EXISTS identite (
            country_iso3 TEXT,
            indicator    TEXT,
            year         INTEGER,
            value        REAL,
            unit         TEXT,
            source       TEXT,
            subcategory  TEXT,
            PRIMARY KEY (country_iso3, indicator, year, subcategory)
        )
    """)
    conn.commit()


def upsert_rows(conn, rows, indicator, subcategory, unit):
    if not rows:
        return 0
    data = [
        (iso3, indicator, year, value, unit, SOURCE, subcategory)
        for (iso3, year, value) in rows
    ]
    conn.executemany("""
        INSERT OR REPLACE INTO identite
            (country_iso3, indicator, year, value, unit, source, subcategory)
        VALUES (?, ?, ?, ?, ?, ?, ?)
    """, data)
    conn.commit()
    return len(data)


# -------------------------------------------------------------
# CONFIGURATION DES ONGLETS
# -------------------------------------------------------------

# (nom_onglet, indicator, subcategory, unit, type_parser)
# type_parser : 'series' | 'snapshot_oil_gas' | 'coal_reserves' | 'mineral'
SHEETS_CONFIG = [
    # Énergie — production (séries temporelles)
    ("Oil Production - barrels",        "energie_production", "petrole",          "kb/j",    "series"),
    ("Gas Production - Bcm",            "energie_production", "gaz",              "Bcm",     "series"),
    ("Coal Production - mt",            "energie_production", "charbon",          "Mt",      "series"),
    ("Nuclear Generation - TWh",        "energie_production", "nucleaire",        "TWh",     "series"),
    ("Solar Generation - TWh",          "energie_production", "solaire",          "TWh",     "series"),
    ("Wind Generation - TWh",           "energie_production", "eolien",           "TWh",     "series"),
    ("Hydro Generation - TWh",          "energie_production", "hydro",            "TWh",     "series"),
    ("Geo Biomass Other - TWh",         "energie_production", "geo_biomasse",     "TWh",     "series"),
    ("Renewable Power (inc hydro)-TWh", "energie_production", "renouvelables_total", "TWh",  "series"),
    # Énergie — réserves (snapshots)
    ("Oil - Proved reserves",           "energie_reserves",   "petrole",          "Mb",      "snapshot_oil_gas"),
    ("Gas - Proved reserves",           "energie_reserves",   "gaz",              "Tcm",     "snapshot_oil_gas"),
    ("Coal - Reserves",                 "energie_reserves",   "charbon",          "Mt",      "coal_reserves"),
    # Minéraux (production + réserves dans le même onglet)
    ("Cobalt P-R",              "mineraux_production", "cobalt",        "kt", "mineral"),
    ("Lithium P-R",             "mineraux_production", "lithium",       "kt", "mineral"),
    ("Copper P-R",              "mineraux_production", "cuivre",        "kt", "mineral"),
    ("Nickel P-R",              "mineraux_production", "nickel",        "kt", "mineral"),
    ("Zinc P-R",                "mineraux_production", "zinc",          "kt", "mineral"),
    ("Bauxite P-R",             "mineraux_production", "bauxite",       "kt", "mineral"),
    ("Aluminium P-R",           "mineraux_production", "aluminium",     "kt", "mineral"),
    ("Natural Graphite P-R",    "mineraux_production", "graphite",      "kt", "mineral"),
    ("Rare Earth metals P-R",   "mineraux_production", "terres_rares",  "kt", "mineral"),
    ("Manganese P-R",           "mineraux_production", "manganese",     "kt", "mineral"),
    ("Tin P-R",                 "mineraux_production", "etain",         "kt", "mineral"),
    ("Vanadium P-R",            "mineraux_production", "vanadium",      "kt", "mineral"),
    ("Platinum Group Metals P-R","mineraux_production", "pgm",          "t",  "mineral"),
]


# -------------------------------------------------------------
# POINT D'ENTRÉE
# -------------------------------------------------------------

def run(filepath=None):
    """
    filepath : chemin vers le fichier EI .xlsx
    Si non fourni, cherche dans etl/sources/uploads/
    """
    if not filepath:
        candidates = [
            os.path.join(os.path.dirname(__file__), "uploads", "EI-Stats-Review-ALL-data.xlsx"),
            os.path.join(os.path.dirname(__file__), "uploads", "EI-Stats-Review.xlsx"),
            os.path.join(os.path.dirname(__file__), "EI-Stats-Review-ALL-data.xlsx"),
        ]
        # Cherche aussi tout fichier EI*.xlsx dans uploads/
        uploads_dir = os.path.join(os.path.dirname(__file__), "uploads")
        if os.path.isdir(uploads_dir):
            for f in os.listdir(uploads_dir):
                if f.startswith("EI") and f.endswith(".xlsx"):
                    candidates.insert(0, os.path.join(uploads_dir, f))
        for c in candidates:
            if os.path.exists(c):
                filepath = c
                break

    if not filepath or not os.path.exists(filepath):
        print("⏭️  Fichier Energy Institute non trouvé — import ignoré.")
        print("   Dépose EI-Stats-Review-ALL-data.xlsx dans etl/sources/uploads/")
        return 0

    print("=" * 60)
    print("ETL — Energy Institute Statistical Review")
    print(f"Fichier : {filepath}")
    print(f"Période : {ANNEE_DEBUT} → {ANNEE_FIN}")
    print("=" * 60)

    try:
        import openpyxl
    except ImportError:
        print("❌ openpyxl non installé — pip install openpyxl")
        return 0

    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    conn = sqlite3.connect(PATH_DB)
    ensure_identite_table(conn)

    total = 0

    for sheet_cfg in SHEETS_CONFIG:
        sheet_name, indicator, subcategory, unit, parser_type = sheet_cfg

        if sheet_name not in wb.sheetnames:
            print(f"  ⚠️  Onglet '{sheet_name}' absent du fichier")
            continue

        ws = wb[sheet_name]
        print(f"\n→ {subcategory} ({indicator})")

        try:
            if parser_type == "series":
                rows = parse_series_sheet(ws, indicator, subcategory, unit)
                nb = upsert_rows(conn, rows, indicator, subcategory, unit)
                total += nb
                print(f"  ✅ {nb} lignes")

            elif parser_type == "snapshot_oil_gas":
                rows = parse_reserves_snapshot(ws, indicator, subcategory, unit)
                nb = upsert_rows(conn, rows, indicator, subcategory, unit)
                total += nb
                print(f"  ✅ {nb} lignes (snapshot)")

            elif parser_type == "coal_reserves":
                rows = parse_coal_reserves(ws, indicator, subcategory, unit)
                nb = upsert_rows(conn, rows, indicator, subcategory, unit)
                total += nb
                print(f"  ✅ {nb} lignes (snapshot)")

            elif parser_type == "mineral":
                prod_rows, res_rows = parse_mineral_sheet(ws, subcategory)
                nb_prod = upsert_rows(conn, prod_rows, "mineraux_production", subcategory, unit)
                nb_res  = upsert_rows(conn, res_rows,  "mineraux_reserves",   subcategory, unit)
                total += nb_prod + nb_res
                print(f"  ✅ production: {nb_prod} lignes | réserves: {nb_res} lignes")

        except Exception as e:
            print(f"  ❌ Erreur : {e}")
            import traceback
            traceback.print_exc()

    conn.close()
    print(f"\n{'='*60}")
    print(f"Energy Institute terminé — {total} lignes au total")
    print(f"{'='*60}")
    return total


if __name__ == "__main__":
    filepath = sys.argv[1] if len(sys.argv) > 1 else None
    run(filepath)
